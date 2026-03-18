"""
Excel 처리 모듈 (pandas + openpyxl 기반)
- 1단계: pandas 고속 탐지
- 2단계: 결과 컬럼 생성 (탐지여부, 탐지된 단어)
- 3단계: Excel 저장 (output_ 접두사)
- 4단계: openpyxl 조건부 서식 적용 (노란색 행 강조)
"""

import re
import warnings
from pathlib import Path
from typing import Set, Dict

import openpyxl
import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# 컬럼명 상수
COL_DETECTED = "발견여부"
COL_KEYWORDS = "발견된 단어"

# 조건부 서식 노란색
YELLOW_FILL = PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
)


# ──────────────────────────────────────────────
# 1단계: pandas 탐지
# ──────────────────────────────────────────────

def _detect_keywords_in_row(row: pd.Series, pattern: re.Pattern) -> Set[str]:
    """한 행에서 발견된 단어 집합을 반환한다."""
    found: Set[str] = set()
    for cell in row:
        if cell is None:
            continue
        for match in pattern.finditer(str(cell)):
            found.add(match.group())
    return found


def _build_pattern(keywords: Set[str]) -> re.Pattern:
    """키워드 set → 정규식 패턴 (대소문자 구분 유지)"""
    escaped = sorted(map(re.escape, keywords), key=len, reverse=True)
    return re.compile("|".join(escaped))


# ──────────────────────────────────────────────
# 2단계: 컬럼 생성
# ──────────────────────────────────────────────

def _add_result_columns(df: pd.DataFrame, pattern: re.Pattern) -> pd.DataFrame:
    """탐지여부 / 발견된 단어 컬럼을 추가(또는 덮어쓰기)한다."""
    for col in (COL_DETECTED, COL_KEYWORDS):
        if col in df.columns:
            df.drop(columns=[col], inplace=True)

    results = df.apply(
        lambda row: _detect_keywords_in_row(row, pattern), axis=1
    )

    df[COL_DETECTED] = results.apply(lambda s: "TRUE" if s else "FALSE")
    df[COL_KEYWORDS] = results.apply(
        lambda s: ", ".join(sorted(s)) if s else ""
    )
    return df


# ──────────────────────────────────────────────
# 3단계: Excel 저장
# ──────────────────────────────────────────────

def _save_excel(df: pd.DataFrame, output_path: Path) -> None:
    df.to_excel(str(output_path), index=False, engine="openpyxl")


# ──────────────────────────────────────────────
# 4단계: openpyxl 조건부 서식
# ──────────────────────────────────────────────

def _apply_conditional_format(output_path: Path) -> None:
    """
    탐지여부 == TRUE 인 행 전체에 노란색 배경 적용.
    wb.close() 호출 금지 (일부 버전 크래시 원인)
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(str(output_path))

    ws = wb.active
    header_row = [cell.value for cell in ws[1]]

    try:
        detected_col_idx = header_row.index(COL_DETECTED) + 1
    except ValueError:
        wb.save(str(output_path))
        return

    detected_col_letter = get_column_letter(detected_col_idx)
    max_row = ws.max_row
    max_col = ws.max_column

    if max_row < 2:
        wb.save(str(output_path))
        return

    last_col_letter = get_column_letter(max_col)
    apply_range = f"A2:{last_col_letter}{max_row}"
    formula = f'=${detected_col_letter}2="TRUE"'

    rule = FormulaRule(formula=[formula], fill=YELLOW_FILL)
    ws.conditional_formatting.add(apply_range, rule)

    wb.save(str(output_path))
    # wb.close() 는 의도적으로 호출하지 않음 (openpyxl 버전 호환성)


# ──────────────────────────────────────────────
# 메인 함수
# ──────────────────────────────────────────────

def process_excel(input_path: str, keywords: Set[str]) -> Dict:
    """
    Excel 파일에서 금지어를 탐지하고 결과를 output_ 파일로 저장한다.
    """
    src = Path(input_path)
    if not src.exists():
        raise FileNotFoundError(f"파일 없음: {input_path}")

    if not keywords:
        raise ValueError("검색 대상 단어 목록이 비어 있습니다.")

    output_path = src.parent / f"output_{src.name}"

    # 1단계: pandas 고속 로딩 (openpyxl 경고 억제)
    # 1단계: pandas 로딩 시 header=None 추가
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        # 💡 header=None: 첫 줄을 헤더로 쓰지 않고 0, 1, 2... 숫자로 이름을 붙임
        df = pd.read_excel(str(src), dtype=str, engine="openpyxl", header=None)

    # 💡 컬럼 개수만큼 A, B, C... 문자로 변환하여 지정
    df.columns = [get_column_letter(i + 1) for i in range(len(df.columns))]

    df.fillna("", inplace=True)

    # 2단계: 탐지 + 컬럼 추가
    pattern = _build_pattern(keywords)
    df = _add_result_columns(df, pattern)

    # 3단계: 저장
    _save_excel(df, output_path)

    # 4단계: 조건부 서식
    _apply_conditional_format(output_path)

    detected_count = int((df[COL_DETECTED] == "TRUE").sum())

    return {
        "output_path": str(output_path),
        "total_rows": len(df),
        "detected_rows": detected_count,
    }
